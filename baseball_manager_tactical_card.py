from ppadb.client import Client
import pyautogui
import time
import sys
from PIL import Image
import os
import hashlib
from collections import Counter
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from collections import defaultdict
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import tkinter as tk
import keyboard
#baseball_manager_tactical_card.py

# GUI
# 변수 정의(사용자에게 입력값을 받을거임)
user_input_value = 0  # 초기값은 0 또는 다른 적절한 값

# 함수: 사용자 입력값을 처리하는 역할
def process_input():
    global user_input_value  # 전역 변수로 사용자 입력값을 저장하기 위해 선언
    user_input = entry.get()  # 사용자가 입력한 값을 가져옴
    try:
        user_input_value = int(user_input)  # 문자열을 정수로 변환하고 변수에 저장
        result_label.config(text=f'입력값 (정수): {user_input_value}')
    except ValueError:
        result_label.config(text="올바른 정수를 입력하세요.")

# 메인 윈도우 생성
window = tk.Tk()
window.title("사용자 입력")

# 레이블 생성
label = tk.Label(window, text="뽑을 수(5배수로 부탁드립니다.) :")
label.pack()

# 입력 필드 생성
entry = tk.Entry(window)
entry.pack()

# 버튼 생성
submit_button = tk.Button(window, text="확인", command=process_input)
submit_button.pack()

# 결과 표시 레이블
result_label = tk.Label(window, text="")
result_label.pack()

# GUI 루프 시작
window.mainloop()
time.sleep(3)

start=time.time() #시간 카운트 시작(툴 사용 시간 알 수 있도록)

# 이미지 파일(ok)의 상대 경로
image_path = "./ok.PNG"

# random_capture 상대경로
random_capture_path = "./random_capture"

stop_requested = False  # 반복문을 종료하기 위한 플래그

def stop_loop(e):
    global stop_requested
    if e.name == 'space':  # 스페이스바를 멈추기 키로 설정
        stop_requested = True

# 특정 키를 감지하는 훅을 추가
keyboard.on_press(stop_loop)

# 가챠

# 이미지 10초안에 안보이면 종료(여기서 말하는 이미지는 ok.png)
def click_image(image_path, confidence, timeout=10):
    start_time = time.time()
    while time.time() - start_time < timeout:
        image_location = pyautogui.locateCenterOnScreen(image_path, confidence=confidence)
        if image_location is not None:
            pyautogui.click(image_location)
            return True
        time.sleep(1)
    return False

time.sleep(2)

#가챠시작
x = 0
max_capture = user_input_value # 사용자 입력값

screen_width, screen_height = pyautogui.size()

# 화면 가로 너비를 7등분, 세로 높이를 3등분
part_width = screen_width // 7
part_height = screen_height // 3
part_height1 = part_height * 3 // 4

x1 = (screen_width // 2) - (part_width // 2)
y1 = (screen_height // 2) - (part_height // 2)

region = (x1, y1, part_width, part_height1)

while x < max_capture:
    if stop_requested:
        break  # 반복문을 멈춥니다.
    #화면 중앙클릭(전술카드가 가운데에 위치함)
    screen_width, screen_height = pyautogui.size()
    center_x = screen_width // 2
    center_y = screen_height // 2
    pyautogui.click(center_x, center_y)

    time.sleep(2)

    if not click_image(image_path, confidence=0.6):
        print("이미지를 찾지 못했습니다.")
        break
    
    time.sleep(1)
    # 한번 전술카드 클릭할때 5뽑하니까 반복문을 넣어줌
    for _ in range(5):
        # 스페이스바를 누르기 전까지 반복합니다.
        if stop_requested:
            break  # 반복문을 멈춥니다.
        time.sleep(6.5)
        # 파일 경로 생성
        screenshot_path = os.path.join(random_capture_path, str(x + 1) + '.png')
    
        capture = pyautogui.screenshot(screenshot_path, region=region)

        if not click_image(image_path, confidence=0.6):
            print("이미지를 찾지 못했습니다.")
            break
        x+=1
        
        
    time.sleep(1.5)

    if not click_image(image_path, confidence=0.6):
        print("이미지를 찾지 못했습니다.")
        break
    time.sleep(3)
    
    if stop_requested:
        break  # 반복문을 멈춥니다.

# 특정 키를 감지하는 훅을 제거
keyboard.unhook_all()

# 가챠 끝내고 데이터 엑셀파일로 저장 시작

# 이미지 파일이 있는 디렉토리 경로
image_dir = random_capture_path

# 이미지 파일의 해시값을 계산하는 함수
def calculate_hash(file_path):
    sha256 = hashlib.sha256()
    with open(file_path, 'rb') as f:
        while True:
            data = f.read(65536)  # 64 KB 단위로 읽어서 해시 계산
            if not data:
                break
            sha256.update(data)
    return sha256.hexdigest()

# 이미지 디렉토리 내의 모든 이미지 파일에 대한 해시값 계산
hash_counter = defaultdict(list)  # 이미지 해시값과 파일 경로를 리스트로 저장
total_image_count = 0


for root, dirs, files in os.walk(image_dir):
    for file in files:
        if file.lower().endswith(('.jpg', '.jpeg', '.png', '.gif')):
            file_path = os.path.join(root, file)
            file_hash = calculate_hash(file_path)
            hash_counter[file_hash].append(file_path)  # 동일한 해시값의 파일 경로를 리스트로 저장
            total_image_count += 1

# 결과를 엑셀 파일에 저장
workbook = openpyxl.Workbook()
worksheet = workbook.active

#헤더
worksheet['A1'] = '인덱스'
worksheet['B1'] = '해시값'
worksheet['C1'] = '이미지'
worksheet['D1'] = '개수'
worksheet['E1'] = '확률'
worksheet['F1'] = '기준값(확률)'

current_row = 2
total_count_d = 0
total_count_e = 0

for index, (hash_value, file_paths) in enumerate(hash_counter.items(), start=1):
    count = len(file_paths)
    probability = count / total_image_count
    worksheet[f'A{current_row}'] = index
    worksheet[f'B{current_row}'] = hash_value
    worksheet[f'D{current_row}'] = count
    worksheet[f'E{current_row}'] = probability
    

    # 이미지를 시트에 추가 (여러 이미지 파일이 있는 경우, 첫 번째 파일만 추가)
    img = Image(file_paths[0])

    # 이미지 크기 조절
    img.width = 80  # 이미지 너비 조절
    img.height = 80  # 이미지 높이 조절

    worksheet.add_image(img, f'C{current_row}')

    current_row += 1
    total_count_d += count
    total_count_e += probability

# D열과 E열의 마지막 셀에 데이터 합 추가
worksheet[f'D{current_row}'] = total_count_d
worksheet[f'E{current_row}'] = total_count_e

# E 열의 각 셀을 백분율 형식으로 표시 (소수점 2자리까지 표시)
for row in worksheet.iter_rows(min_row=2, max_row=current_row, min_col=5, max_col=5):
    for cell in row:
        if True:
            percentage_value = cell.value * 100
            cell.value = f"{percentage_value:.2f}%"  # 소수점 2자리까지 표시
            cell.number_format = '0.00%'  # 백분율 형식으로 표시

# 데이터 read편하게 css

column_letter = get_column_letter(3)  # C열을 선택
worksheet.column_dimensions[column_letter].width = 10  # C열 너비 조절

column_letter = get_column_letter(6)  # E열을 선택
worksheet.column_dimensions[column_letter].width = 20  # E열 너비 조절

row_height = 60  # 행의 높이 조절
for i in range(2, current_row):
    worksheet.row_dimensions[i].height = row_height


# 모든 데이터 가운데 정렬
for row in worksheet.iter_rows(min_row=1, max_row=current_row, min_col=1, max_col=6):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# 엑셀 헤더 부분의 텍스트를 볼드 처리
header_cells = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1']  # 헤더 셀의 위치
for cell in header_cells:
    worksheet[cell].font = Font(bold=True)

# B 열을 숨깁니다(해시값은 이미지를 불러오기위한 수단이기때문에 숨김처리해줌)
column_letter = get_column_letter(2)  # B열을 선택
worksheet.column_dimensions[column_letter].hidden = True

# 헤더 행 (A1:F1)을 틀고정
worksheet.freeze_panes = "A2"

excel_file = "./xl/전술카드 확률표.xlsx"       
workbook.save(excel_file)
print(f'캡쳐 이미지 총 개수는 {total_count_d}개 입니다.')
print(f'결과가 {excel_file}에 저장되었습니다.')
print(f"{time.time()-start:.4f} sec") # 종료와 함께 수행시간 출력

